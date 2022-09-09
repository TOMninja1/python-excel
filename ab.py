
import pandas as pd
import openpyxl 
pf=pd.read_excel('EXCEL/e-folder/a1.xlsx')
print(pf)
wa=pf[pf["店舗名"]=="和食 清風"]
print(wa)
# print(pf[pf["店舗名"]=="オステリアSeifu "])
wb=pf[pf["店舗名"]=="Bar Seifu"]
wc=pf[pf["店舗名"]=="オステリアSeifu"]

# print(type(wa))
# print(wb)
# wb=openpyxl.Workbook()
# wb.create_sheet(title='和食 清風')
# wb.create_sheet(title='オステリアseifu')
# wb.create_sheet(title='bar sebifu')
# print(wb.sheetnames)
# wb.save('EXCEL/auto1.xlsx')
with pd.ExcelWriter('EXCEl/auto1.xlsx') as writer:
    wa.to_excel(writer, sheet_name='和食 清風')
    wb.to_excel(writer, sheet_name='Bar Seifu')
    wc.to_excel(writer, sheet_name='オステリアSeifu')
    # pf.to_excel(writer, sheet_name='Bar Seifu ')

#pf[pf["商品名"] == "茶碗蒸し"].sum()
wa.query("商品名=='茶碗蒸し'").sum()
#df[['氏名','売上金額']].groupby('氏名').sum()
print(wa)


